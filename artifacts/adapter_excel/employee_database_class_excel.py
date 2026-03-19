import pandas as pd
import openpyxl
from openpyxl.workbook import Workbook
import os


class EmployeeDatabase:
    def __init__(self, file_path="employee_records_DB.xlsx"):
        self.file_path = file_path
        if os.path.exists(self.file_path):
            self.workbook = openpyxl.load_workbook(self.file_path)
            self.sheet = self.workbook.active
        else:
            self.workbook = openpyxl.Workbook()
            self.sheet = self.workbook.active
            self.workbook.save(self.file_path)

    def __del__(self):
        self.workbook.close()

    def readSheet(self):
        return pd.read_excel(self.file_path, sheet_name="Sheet")

    def createHead(self):
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb["Sheet"]
        ws.append([
            "id",
            "first_name",
            "last_name",
            "birthdate",
            "hire_date",
            "job_title"
        ])
        wb.save(self.file_path)

    def insert_employee(self, first_name, last_name, birthdate, hire_date, job_title):
        if self.sheet.cell(1, 1).value is None:
            self.createHead()
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb["Sheet"]
        idNum = ws.max_row
        ws.append([
            idNum,
            first_name,
            last_name,
            birthdate,
            hire_date,
            job_title,
        ])
        wb.save(self.file_path)

    def get_employee(self, first_name, last_name):
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb["Sheet"]
        for row in ws.iter_rows(min_row=2):
            firstN = row[1].value
            lastN = row[2].value
            if firstN == first_name and lastN == last_name:
                return (f"id: {row[0].value}, "
                        f"Name: {row[1].value} {row[2].value}, "
                        f"Birthday: {row[3].value}, "
                        f"Hire Date: {row[4].value}, "
                        f"Job Title: {row[5].value}")

        return None

    def getAllEmp(self):
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb["Sheet"]
        employees =[]
        for row in ws.iter_rows(min_row=2):
            employees.append(f"id: {row[0].value}, "
                    f"Name: {row[1].value} {row[2].value}, "
                    f"Birthday: {row[3].value}, "
                    f"Hire Date: {row[4].value}, "
                    f"Job Title: {row[5].value}")
        return employees

    def close(self):
        self.workbook.close()
